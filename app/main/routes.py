from time import strptime
from flask import render_template, url_for, request, redirect, flash, jsonify, current_app
from flask.helpers import send_from_directory
from flask_login import current_user, login_required
from datetime import datetime
from app import db
from app.main import bp
from app.models import *
from app.main.forms import *
from app.utils import *
from wtforms.fields.core import BooleanField
import logging, openpyxl, os
from openpyxl.styles import Alignment, Font, Border, Side

# Import the logger assigned to the application
logger = logging.getLogger(__name__)

@bp.route('/', methods=['GET', 'POST'])
@bp.route('/index', methods=['GET', 'POST'])
def index():

    summary = {}

    users = User.query.all()
    summary["users"] = users

    companies = Company.query.all()
    summary["companies"] = companies

    projects = Project.query.all()
    summary["projects"] = projects

    jobs = Job.query.all()
    summary["jobs"] = jobs

    groups = Group.query.all()
    summary["groups"] = groups

    channels = Channel.query.all()
    summary["channels"] = channels

    test_equipment = TestEquipment.query.all()
    summary["test_equipment"] = test_equipment

    calibration_records = CalibrationRecord.query.all()
    summary["calibration_records"] = calibration_records

    logger.debug(f'Projects = {summary["projects"]}')

    return render_template('index.html', title='Home', summary=summary)


@bp.route('/user/<username>')
@login_required
def user(username):

    # Get the profile of the current_user
    user = User.query.filter_by(username=username).first_or_404()
    return render_template('user.html', user=user)


@bp.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():

    form = EditProfileForm(current_user.username)

    # User has changed their profile information
    if form.validate_on_submit():

        # Update the database with the user's changes to their profile information
        current_user.username = form.username.data
        db.session.commit()
        flash('Your changes have been saved.')
        return redirect(url_for('main.edit_profile'))

    # User is only visiting their profile
    elif request.method == 'GET':
        form.username.data = current_user.username

    return render_template('edit_profile.html', title='Edit Profile', form=form)


@bp.route('/add_company', methods=['GET', 'POST'])
@login_required
def add_company():

    # Create the form
    form = AddCompanyForm()

    # User has added a new client
    if form.validate_on_submit():

        # Extract the Company's category from the form
        company = Company(
            name=form.name.data,
            category=form.category.data
        )

        db.session.add(company)
        db.session.commit()
        flash(f'Company "{company.name}" has been added to the database.')

        return redirect(url_for('main.index'))

    return render_template('add_item.html', title='Add Company', form=form, item='Company')


@bp.route('/add_project', methods=['GET', 'POST'])
@login_required
def add_project():

    # Create the form
    form = AddProjectForm()

    # Generate the list of SelectField choices to populate in the form
    form.client.choices = EMPTY_SELECT_CHOICE + [(c.id, c.name) for c in Company.clients()]
    form.supplier.choices = EMPTY_SELECT_CHOICE + [(s.id, s.name) for s in Company.suppliers()]

    # User has added a new project
    if form.validate_on_submit():

        # Add the new project to the database
        project = Project(
            name=form.name.data,
            number=form.number.data
        )
        db.session.add(project)
        db.session.commit()

        # Add the specified Companies to the project
        client = Company.query.filter_by(id=form.client.data).first()
        supplier = Company.query.filter_by(id=form.supplier.data).first()
        project.add_company(client)
        project.add_company(supplier)
        db.session.commit()

        flash(f'Project "{project.name}" has been added to the database.')

        return redirect(url_for('main.projects'))

    return render_template('add_item.html', title='Add Project', form=form, item='Project')


@bp.route('/projects', methods=['GET', 'POST'])
@login_required
def projects():

    projects = Project.query.all()

    return render_template('projects.html', title='Projects List', projects=projects)


@bp.route('/projects/<project_id>/add_job', methods=['GET', 'POST'])
@login_required
def add_job(project_id):

    # Get the associated project information
    project = Project.query.filter_by(id=project_id).first()
    client = project.client()
    supplier = project.supplier()

    # Generate the choices lists for the form's SelectFields
    CLIENT_NAME_CHOICES = [(client.id, client.name)]
    SUPPLIER_NAME_CHOICES = [(supplier.id, supplier.name)]
    PROJECT_NUMBER_CHOICES = [(project_id, project.number)]
    PROJECT_NAME_CHOICES = [(project_id, project.name)]

    # Create and prepopulate the form
    form = AddJobForm(
        client_name=client.id,
        project_number=project_id,
        project_name=project_id
    )

    # Add the project_id from which the new job was requested by the user to be added to
    form.project_id = project_id

    # Assign the SelectField choices to populate in the form
    form.client_name.choices = CLIENT_NAME_CHOICES
    form.supplier_name.choices = SUPPLIER_NAME_CHOICES
    form.project_number.choices = PROJECT_NUMBER_CHOICES
    form.project_name.choices = PROJECT_NAME_CHOICES

    # User has added a new job
    if form.validate_on_submit():
        # Add the new job to the database
        job = Job(
            project_id=form.project_name.data,
            name=form.name.data,
            stage=form.stage.data,
            phase=form.phase.data            
        )
        db.session.add(job)
        db.session.commit()
        flash(f'Job "{job.stage} {job.phase}" has been added to the {job.project.name} project.')
        
        return redirect(url_for('main.jobs', project_id=project_id))

    return render_template('add_item.html', title='Add Job', form=form, item='Job')


@bp.route('/projects/<project_id>/jobs', methods=['GET', 'POST'])
@login_required
def jobs(project_id):

    project = Project.query.filter_by(id=project_id).first()
    jobs = Job.query.filter_by(project_id=project_id).all()

    return render_template('jobs.html', title='Job List', jobs=jobs, project=project)


@bp.route('/job/<job_id>/add_group', methods=['GET', 'POST'])
@login_required
def add_group(job_id):

    # Create the form
    form = AddGroupForm()

    # Add the job_id from which the new group was requested by the user to be added to
    form.job_id.data = job_id

    # User has added a new group to an existing job
    if form.validate_on_submit():

        # Add the new group to the database
        group = Group(
            name=form.name.data,
            job_id=form.job_id.data
        )
        db.session.add(group)
        db.session.commit()
        flash(f'Group "{group.name}" has been added to the {group.job.stage} {group.job.phase} job for the {group.job.project.name} project.')

        return redirect(url_for('main.groups', job_id=job_id))

    return render_template('add_item.html', title='New Group', form=form, item="Group")

@bp.route('/job/<job_id>/groups', methods=['GET', 'POST'])
@login_required
def groups(job_id):

    job = Job.query.filter_by(id=job_id).first()
    groups = Group.query.filter_by(job_id=job_id).all()

    return render_template('groups.html', title='Group List', groups=groups, job=job)


@bp.route('/group/<group_id>/add_channel', methods=['GET', 'POST'])
@login_required
def add_channel(group_id):

    # Pre-load checkbox-style buttons for each TestEquipmentType into the form
    test_equipment_types = [test_equipment for test_equipment in TestEquipmentType.query.all()]
    for test_equipment_type in test_equipment_types:
        id = 'checkbox-' + (test_equipment_type.name).replace('_', '-')
        setattr(AddChannelForm, f'checkbox_{test_equipment_type.name}',
            BooleanField(id=id, render_kw=CUSTOM_FORM_CLASS))

    # Create the form
    form = AddChannelForm()

    # Add the group_id from which the new channel was requested by the user to be added to
    form.group_id.data = group_id
    has_full_scale_range = True

    if request.method == 'POST' and form.full_scale_range.data == None:
        # Remove the Full Scale Range field off the form if the ErrorType is not %FS
        del form.full_scale_range
        has_full_scale_range = False

    if form.validate_on_submit():

        '''Extract the form data'''
        # Basic Channel Info
        group_id = form.group_id.data
        base_name = form.name.data
        suffix = int(form.suffix.data)
        quantity = int(form.quantity.data)

        # Measurement Info
        measurement_type = form.measurement_type.data
        measurement_units = form.measurement_units.data
        min_range = form.min_range.data
        max_range = form.max_range.data
        full_scale_range = form.full_scale_range.data if has_full_scale_range else None

        # Tolerance Info
        max_error = form.max_error.data
        error_type = form.error_type.data

        # Signal Injection Info
        min_injection_range = form.min_injection_range.data
        max_injection_range = form.max_injection_range.data
        injection_units = form.injection_units.data

        # TestPoint Info
        testpoint_list_type = form.testpoint_list_type.data
        testpoint_list_data = form.testpoint_list.data
        num_testpoints = int(form.num_testpoints.data)

        # Extract the list of custom TestPoints if selected in the form
        injection_values = []
        test_values = []
        if testpoint_list_type == TestPointListType.CUSTOM:
            for i, values in enumerate(testpoint_list_data):
                injection_values.append(values["injection_value"])
                test_values.append(values["test_value"])

        # Extract the required TestEquipmentTypes selected in the form
        required_test_equipment_types = []
        for test_equipment_type in test_equipment_types:
            if form.data['checkbox_' + test_equipment_type.name]:
                required_test_equipment_types.append(test_equipment_type)
                
        # Setup variables to prep for the iteration through channel creation
        start = suffix
        end = suffix + quantity

        for i in range(start, end):

            # Create the new name for the channel
            name = base_name + f'{i:03d}'

            # Create the new channel and add it to the database
            channel = Channel(
                name=name,
                group_id=group_id,
                measurement_type=measurement_type,
                measurement_units=measurement_units,
                min_range=min_range,
                max_range=max_range,
                full_scale_range=full_scale_range,
                max_error=max_error,
                error_type=error_type,
                min_injection_range=min_injection_range,
                max_injection_range=max_injection_range,
                injection_units=injection_units,
            )
            db.session.add(channel)
            db.session.commit()            

            # Build the list of TestPoints for the new channel
            channel.build_testpoint_list(num_testpoints, testpoint_list_type,
                injection_values, test_values)

            # Add the required TestEquipmentType for the new channel
            for test_equipment_type in required_test_equipment_types:
                channel.add_test_equipment_type(test_equipment_type)

            # Add the required Approvals for the new channel
            channel.update_required_approvals()
            db.session.commit()
        
        flash(f'{quantity} new channels have been added to the {channel.group.name} group each with {num_testpoints} testpoints.')
        
        return redirect(url_for(f'main.channels', group_id=group_id))

    # Display the errors if there are any
    if len(form.errors.items()) > 0:
        print(form.errors.items())

        # dict_items([('min_range', ['Not a valid float value'])])
        # Clear the form and flash a message
        message = 'Form reset due to the following errors:'
        for error in form.errors.items():
            message = message + f'\n{error[0]}: {error[1][0]}, '
        
        flash(message)

        return redirect(url_for('main.add_channel', group_id=group_id))

    return render_template('add_channel.html', title='Add Channel', form=form, test_equipment_types=test_equipment_types)


@bp.route('/group/<group_id>/channels', methods=['GET', 'POST'])
@login_required
def channels(group_id):

    # Get a list of all the specified Group's Channels
    group = Group.query.filter_by(id=group_id).first()
    channels = Channel.query.filter_by(group_id=group_id).order_by('name').all()

    # Create the master form to add each channel_form into
    channels_form = ChannelsForm()

    for channel in channels:

        # Get a list of all the TestPoints in each Channel
        testpoints = channel.testpoints

        # Create the channel_form to add each testpoint_form into
        channel_form = ChannelForm()
        channel_form.notes.data = channel.notes

        for testpoint in testpoints:

            # Create the testpoint_form for each testpoint
            testpoint_form = TestPointForm()

            # Add the testpoint_form into the channel_form
            channel_form.testpoints.append_entry(testpoint_form)
        
        # Add the channel_form into the master channels_form
        channels_form.channels.append_entry(channel_form)

    return render_template('channels.html', title='Channel List', channels=channels,
        channels_form=channels_form, group=group, timestamp=datetime.now())


@bp.route('/delete_channel', methods=['POST'])
@login_required
def delete_channel():

    # Extract the request's form dictionary
    data = request.form.to_dict()
    print(data)
    CHANNEL_ID = 'channel_id'

    if CHANNEL_ID in data:

        # Get the channel from the ajax request
        channel = Channel.query.filter_by(id=data[CHANNEL_ID]).first()

        # Delete the channel and all its dependencies
        channel.delete_all_records()
        db.session.delete(channel)
        db.session.commit()
    else:
        raise ValueError(f'{CHANNEL_ID} not found in ajax request:\n{data}')
        
    response = {
        'message': f'{channel} has been successfully deleted.'
    }
    return jsonify(response)

@bp.route('/delete_testpoint', methods=['POST'])
@login_required
def delete_testpoint():

    # Extract the request's form dictionary
    data = request.form.to_dict()
    print(data)
    CHANNEL_ID = 'channel_id'
    TESTPOINT_ID = 'testpoint_id'

    if CHANNEL_ID in data and TESTPOINT_ID in data:

        # Get the Channel and TestPoint from the ajax request
        channel = Channel.query.filter_by(id=data[CHANNEL_ID]).first()
        testpoint = TestPoint.query.filter_by(id=data[TESTPOINT_ID]).first()

        # Delete the TestPoint from the Channel
        channel.remove_testpoint(testpoint)
        db.session.commit()

        # Update the Channel and its parent items
        last_updated = datetime.utcnow()
        channel.last_updated = last_updated
        channel.update_each_parent_status(last_updated)

        response = {
            "message": f'TestPoint for {channel} has been successfully deleted.',
            "last_updated": last_updated,
            "num_testpoints": channel.num_testpoints(),
            "progress": channel.testpoint_progress(),
            "num_passed": channel.testpoint_stats()[TestResult.PASS.value],
            "status": channel.status
        }
        return jsonify(response)

    else:
        raise ValueError(f'{CHANNEL_ID} or {TESTPOINT_ID} not found in ajax request:\n{data}')


@bp.route('/update_channel', methods=['POST'])
@login_required
def update_channel():

    # Channel Field Constants
    CHANNEL_ID = 'channel_id'
    INTERFACE = 'interface'
    NOTES = 'notes'
    LAST_UPDATED = 'last_updated'
    TEST_EQUIPMENT_ID = 'test_equipment_id'
    TEST_EQUIPMENT_TYPE_ID = 'test_equipment_type_id'

    # Other constants
    MESSAGE = 'message'
    SUPPLIER_APPROVAL = 'supplier_approval'
    CLIENT_APPROVAL = 'client_approval'
    ADD_TESTPOINT = 'addTestPoint'
    NEW_CHANNEL_NAME = 'newChannelName'
    NEW_INJECTION_UNITS = 'newInjectionUnits'
    NEW_TEST_UNITS = 'newTestUnits'
    NOMINAL_INJECTION_VALUE = 'nominal_injection_value'
    NOMINAL_TEST_VALUE = 'nominal_test_value'

    # Variables to keep track of the updated fields
    updated_fields = []
    last_updated = datetime.utcnow()

    # Extract the request's form dictionary
    data = request.form.to_dict()
    print(data)

    if CHANNEL_ID in data:
        # Get the channel from the ajax request and remove it from the keys to check
        channel = Channel.query.filter_by(id=request.form[CHANNEL_ID]).first()
        data.pop(CHANNEL_ID)
    else:
        raise ValueError(f'{CHANNEL_ID} not found in ajax request:\n{data}')
    
    if TEST_EQUIPMENT_TYPE_ID in data:
        # Get the test_equipment_type_id from the ajax request and remove it from the keys to check
        test_equipment_type_id = data[TEST_EQUIPMENT_TYPE_ID]
        data.pop(TEST_EQUIPMENT_TYPE_ID)

    # Iterate through each field in the request and update the Channel accordingly
    for key, value in data.items():        

        if key == INTERFACE:            
            channel.interface = none_if_empty(value)
        
        if key == NOTES:            
            channel.notes = none_if_empty(value)

        if key == TEST_EQUIPMENT_ID:

            # Get the current TestEquipment assigned to the channel
            current_test_equipment = channel.current_test_equipment(test_equipment_type_id)
            has_current_test_equipment = (current_test_equipment is not None)

            # The current TestEquipment is the same as the selected TestEquipment so ignore
            if has_current_test_equipment:
                if current_test_equipment.id == value:
                    continue

            # A new TestEquipment has been selected
            new_test_equipment = TestEquipment.query.filter_by(id=value).first()

            # Create a record of the TestEquipment being used by the channel
            record = ChannelEquipmentRecord(
                channel_id=channel.id,
                test_equipment_id=new_test_equipment.id,
                test_equipment_type_id=test_equipment_type_id,
                timestamp=last_updated,
                calibration_due_date=new_test_equipment.due_date()
            )
            channel.equipment_records.append(record)
            db.session.commit()

        if key == SUPPLIER_APPROVAL or key == CLIENT_APPROVAL:

            # Update the approval status of the channel
            if value == 'true':
                channel.add_approval(current_user)
            else:
                channel.remove_approval(current_user)
            db.session.commit()

        if key == ADD_TESTPOINT:

            # Extract the TestPoint data from the request
            if NOMINAL_INJECTION_VALUE in data:
                nominal_injection_value = data[NOMINAL_INJECTION_VALUE]

            if NOMINAL_TEST_VALUE in data:
                nominal_test_value = data[NOMINAL_TEST_VALUE]

            # TODO: Add a check for if the nominal_injection_value is > the max_range or < min_range, to update

            new_testpoint = TestPoint(
                channel_id = channel.id,
                nominal_injection_value = nominal_injection_value,
                nominal_test_value = nominal_test_value
            )
            channel.add_testpoint(new_testpoint)
            db.session.commit()
        
        if key == NEW_CHANNEL_NAME:
            channel.name = none_if_empty(value)
        
        if key == NEW_INJECTION_UNITS:
            channel.injection_units = value

        if key == NEW_TEST_UNITS:  
            channel.measurement_units = value

        # Add the processed key to the list of updated fields
        updated_fields.append(key)
    
    # Update the status and last_updated time on each parent item
    channel.last_updated = last_updated
    channel.update_each_parent_status(last_updated)

    # Save the changes to the database
    db.session.commit()

    # Load the last_updated time into the json payload for a successful ajax request
    response = {
        MESSAGE: f'{channel} has successfully updated the following fields: {updated_fields}',
        LAST_UPDATED: last_updated
    }
    
    if ADD_TESTPOINT in data:

        # Figure out the position of the new TestPoint in the existing list of TestPoints
        new_position = 0
        for testpoint in channel.testpoints.order_by('nominal_injection_value').all():

            # Finds the first TestPoint with a larger value so that the new TestPoint can be added before it
            if testpoint.nominal_injection_value >= new_testpoint.nominal_injection_value:
                break
            new_position += 1

        # Add the necessary TestPoint data to aid in the new TestPoint being created
        response["testpoint_data"] = {
            "testpoint_id": new_testpoint.id,
            "nominal_injection_value": new_testpoint.nominal_injection_value,
            "nominal_test_value": new_testpoint.nominal_test_value,
            "num_testpoints": channel.num_testpoints(),
            "lower_limit": round(new_testpoint.lower_limit(), 5),
            "upper_limit": round(new_testpoint.upper_limit(), 5),
            "new_position": new_position,
            "progress": channel.testpoint_progress(),
            "num_passed": channel.testpoint_stats()[TestResult.PASS.value],
            "status": channel.status
        }

    return jsonify(response)

@bp.route('/update_testpoint', methods=['POST'])
@login_required
def update_testpoint():
    
    # TestPoint and Channel Field Constants
    CHANNEL_ID = 'channel_id'
    TESTPOINT_ID = 'testpoint_id'
    MEASURED_INJECTION_VALUE = 'measured_injection_value'
    MEASURED_TEST_VALUE = 'measured_test_value'
    MEASURED_ERROR = 'measured_error'
    TEST_RESULT = 'test_result'
    STATUS = 'status'
    LAST_UPDATED = 'last_updated'

    # Other constants
    MESSAGE = 'message'
    PROGRESS = 'progress'
    NUM_PASSED = 'num_passed'

    # Variables to keep track of the updated fields
    updated_fields = []
    num_fields = 0

    # Evaluation variables
    has_channel = False

    # Extract the request's form dictionary
    data = request.form.to_dict()

    # Check which TestPoint is being updated
    if TESTPOINT_ID in data:        

        # Remove the testpoint_id from the data to check how many parameters are being updated
        testpoint = TestPoint.query.filter_by(id=request.form[TESTPOINT_ID]).first()
        data.pop(TESTPOINT_ID)
    else:
        raise ValueError(f'{TESTPOINT_ID} not found in ajax request:\n{data}')
        
    if CHANNEL_ID in data:        

        # Remove the channel_id from the data to check how many parameters are being updated
        channel = Channel.query.filter_by(id=request.form[CHANNEL_ID]).first()
        data.pop(CHANNEL_ID)
        has_channel = True
    else:
        raise ValueError(f'{CHANNEL_ID} not found in ajax request:\n{data}')
    
    # Calculate the new number of variables to iterate over
    num_fields = len(data)

    # Iterate through each field in the request and update the TestPoint accordingly
    for key, value in data.items():

        if key == MEASURED_INJECTION_VALUE:            
            testpoint.measured_injection_value = none_if_empty(value)

        if key == MEASURED_TEST_VALUE:
            testpoint.measured_test_value = none_if_empty(value)

        if key == MEASURED_ERROR:
            testpoint.measured_error = none_if_empty(value)

        if key == TEST_RESULT:
            testpoint.test_result = value

        # Add the processed key to the list of updated fields
        updated_fields.append(key)

    # Check to make sure any fields got updated
    num_updated = len(updated_fields)
    if not num_updated == num_fields:
        raise ValueError(f'Error updating TestPoint fields. Only {num_updated}/{num_fields} updated successfully.')

    # Update the last_updated time now that changes have been made
    last_updated = datetime.utcnow()
    testpoint.last_updated = last_updated

    # Update the status and last_update time of the updated channel, group, job and project
    if has_channel:
        channel.last_updated = last_updated
        channel.update_each_parent_status(last_updated)

    # Save the changes to the database
    db.session.commit()

    # Load the last_updated time into the json payload for a successful ajax request
    response = {
        MESSAGE: f'{testpoint} has successfully updated the following fields: {updated_fields}',
        LAST_UPDATED: last_updated,
        PROGRESS: channel.testpoint_progress(),
        STATUS: channel.status,
        NUM_PASSED: channel.testpoint_stats()[TestResult.PASS.value]
    }

    return jsonify(response)


@bp.route('/get_updated_group_data', methods=['GET', 'POST'])
def get_updated_group_data():

    # Extract the request's form dictionary
    data = request.form.to_dict()
    print(data)

    # Extract the data from the request
    group = Group.query.filter_by(id=data["group_id"]).first()
    page_load_timestamp_raw = data["page_load_timestamp"]
    page_load_timestamp = datetime.strptime(page_load_timestamp_raw, '%Y-%m-%d %H:%M:%S.%f')

    # Dynamically compile a list of items updated by the current_user to exclude them from the query?
    # May not be able to if another user edits a testpoint/channel after the current user has
    # Should maybe add a last_updated_by_user field on each channel and testpoint?

    channels_to_update = []
    testpoints_to_update = []

    # Find all the channels and testpoints which have been updated after the page was loaded
    # Note: This intent is to find all changes from other users that are not visibile to the current_user
    for channel in group.channels.all():

        # Find all the channels which have been updated after the page was loaded
        if channel.last_updated > page_load_timestamp:
            channels_to_update.append(channel.id)
        
        for testpoint in channel.testpoints.all():
            
            # Find all testpoints which have been updated after the page was loaded
            if testpoint.last_updated > page_load_timestamp:
                testpoints_to_update.append(testpoint.id)

                # Get: 

    
    print(f'Channel IDs to update: {channels_to_update}')
    print(f'TestPoint IDs to update: {testpoints_to_update}')

    response = {
        'message': 'Successfully fetched for group data',
        'group': group.name,
        'page_load_timestamp': page_load_timestamp
    }

    return jsonify(response)


@bp.route('/add_test_equipment', methods=['GET', 'POST'])
@login_required
def add_test_equipment():

    # Create the form
    form = AddTestEquipmentForm()

    # User has added a new TestEquipment
    if form.validate_on_submit():

        # Add the new TestEquipment to the database
        test_equipment = TestEquipment(
            name=form.name.data,
            manufacturer=form.manufacturer.data,
            model_num=form.model_num.data,
            serial_num=form.serial_num.data,
            asset_id=form.asset_id.data
            )
        db.session.add(test_equipment)
        db.session.commit()

        # Add the starting calibration information for the TestEquipment
        calibration_record = CalibrationRecord(
            calibration_date=form.calibration_date.data,
            calibration_due_date=form.calibration_due_date.data,
            test_equipment_id=test_equipment.id
        )
        test_equipment.add_calibration_record(calibration_record)
        db.session.commit()

        # Add the new TestEquipment to the list of TestEquipmentTypes 
        test_equipment_type = TestEquipmentType.query.filter_by(name=test_equipment.name).first()
        test_equipment_type.add_test_equipment(test_equipment)
        db.session.commit()

        flash(f'Test Equipment "{test_equipment.asset_id}" {test_equipment.name} has been added to the database.')
        
        return redirect(url_for('main.test_equipment'))

    return render_template('add_item.html', title='Add Test Equipment', form=form, item='Test Equipment')


@bp.route('/test_equipment/<test_equipment_id>/add_calibration_record', methods=['GET', 'POST'])
@login_required
def add_calibration_record(test_equipment_id):

    test_equipment = TestEquipment.query.filter_by(id=test_equipment_id).first()

    # Create the form
    form = AddCalibrationRecordForm()

    # User has added a new CalibrationRecord
    if form.validate_on_submit():

        # Add the new CalibrationRecord to the database
        calibration_record = CalibrationRecord(
            calibration_date=form.calibration_date.data,
            calibration_due_date=form.calibration_due_date.data,
            test_equipment_id=test_equipment_id
        )
        db.session.add(calibration_record)
        db.session.commit()
        flash(f'Calibration Record has been added for "{test_equipment}" with a due date of {calibration_record.calibration_due_date}.')

        return redirect(url_for('main.test_equipment'))

    return render_template('add_item.html', title='Add Calibration Record', form=form, item='Calibration Record')


@bp.route('/test_equipment', methods=['GET', 'POST'])
@login_required
def test_equipment():

    test_equipment = TestEquipment.query.all()

    return render_template('test_equipment.html', title='Test Equipment List', test_equipment=test_equipment)


@bp.route('/projects/<project_id>/edit_project_members', methods=['GET', 'POST'])
@login_required
def edit_project_members(project_id):

    # Get a list of all the users
    users = User.query.order_by('first_name').all()

    # Get a list of the existing members on the Project
    project = Project.query.filter_by(id=project_id).first()

    # TODO: Split up form to 2 sections: Supplier and Client

    for user in users:

        # Gather all the existing members on the project to pre-populate the fields
        member_exists = project.has_member(user)

        # Add the button for the User to the form
        id = 'checkbox-user-' + str(user.id)
        setattr(UpdateProjectForm, f'checkbox_user_{user.id}',
            BooleanField(id=id, render_kw=CUSTOM_FORM_CLASS, default=member_exists)) 
    
    # Create the form
    form = UpdateProjectForm()

    if form.validate_on_submit():
        
        # Result checking variables
        added_users = []
        removed_users = []

        # Update the members assigned to the project
        for user in users:
            if form.data['checkbox_user_' + str(user.id)]:

                # Existing member has not been removed
                if project.has_member(user):
                    continue

                # Add the new member to the project
                project.add_member(user)
                added_users.append(user.username)
            else:
                # Existing member has been removed
                if project.has_member(user):
                    project.remove_member(user)
                    removed_users.append(user.username)                
        
        # Save to the database
        db.session.commit()
        flash('Project Members have been successfully updated.')
        print(f'The following users were added to the {project.name} project: \
            \nAdded Users: {added_users}\nRemoved Users: {removed_users} \
            \nUpdated Project Member List: {project.members}')

        return redirect(url_for('main.projects', project_id=project_id))
    
    # Display the errors if there are any
    if len(form.errors.items()) > 0:
        print(form.errors.items())

    return render_template('edit_project_members.html', title='Edit Project Members',
        form=form, users=users, project=project)
        

@bp.route('/projects/<project_id>/edit_project_test_equipment.html', methods=['GET', 'POST'])
@login_required
def edit_project_test_equipment(project_id):

    # Get a list of the TestEquipmentTypes
    test_equipment_types = TestEquipmentType.query.all()
    
    # Get a list of the existing TestEquipment that has been assigned to the Project
    project = Project.query.filter_by(id=project_id).first()

    for test_equipment_type in test_equipment_types:        
        for test_equipment in test_equipment_type.test_equipment:

            # Get all the existing equipment on the project to pre-populate the fields
            equipment_exists = project.has_test_equipment(test_equipment)

            # Add a button for the TestEquipment to the form
            id = 'checkbox-equipment-' + str(test_equipment.id)
            setattr(UpdateProjectForm, f'checkbox_equipment_{test_equipment.id}',
                BooleanField(id=id, render_kw=CUSTOM_FORM_CLASS, default=equipment_exists))

    # Create the form
    form = UpdateProjectForm()

    if form.validate_on_submit():

        # Result checking variables
        added_equipment = []
        removed_equipment = []

        # Update the test equipment assigned to the project
        for test_equipment_type in test_equipment_types:
            for test_equipment in test_equipment_type.test_equipment:

                if form.data['checkbox_equipment_' + str(test_equipment.id)]:

                    # Existing test equipment has not been removed
                    if project.has_test_equipment(test_equipment):
                        continue

                    # Add the new test equipment to the project
                    project.add_test_equipment(test_equipment)
                    added_equipment.append(test_equipment.name)
                else:
                    # Existing test equipment has been removed
                    if project.has_test_equipment(test_equipment):
                        project.remove_test_equipment(test_equipment)
                        removed_equipment.append(test_equipment.name)
        
        # Save to the database
        db.session.commit()
        flash('Project Equipment List has been successfully updated.')
        print(f'The following test equipment was added to the {project.name} project: \
            \nAdded Equipment: {added_equipment}\nRemoved Equipment: {removed_equipment} \
            \nUpdated Project Equipment List: {project.test_equipment.all()}')

        return redirect(url_for('main.projects', project_id=project_id))

    # Display the errors if there are any
    if len(form.errors.items()) > 0:
        print(form.errors.items())

    return render_template('edit_project_test_equipment.html', title='Edit Project Test Equipment',
        form=form, test_equipment_types=test_equipment_types, project=project)


@bp.route('/test', methods=['GET', 'POST'])
def test():

    project = Project.query.first()
    test_equipment_types = TestEquipmentType.query.all()

    return render_template('project.html', title='Test Project', project=project,
        test_equipment_types=test_equipment_types)

@bp.route('/test2', methods=['GET', 'POST'])
def test2():

    return render_template('test2.html', title="Test Items")

@bp.route('/generate_channel_report/<job_id>', methods=['GET', 'POST'])
def generate_channel_report(job_id):

    def apply_horizontal_border(sheet, border_type, border_style, row_num, start_col, end_col):

        # Define the border formats
        side = Side(border_style=border_style, color="000000")        
        if border_type == TOP_BORDER_TYPE:
            border = Border(top=side)
        elif border_type == BOTTOM_BORDER_TYPE:
            border = Border(bottom=side)

        # Apply the border
        # Note: Adding +1 due to starting at index 1
        for i in range(end_col - start_col + 1):
            sheet.cell(row=row_num, column=start_col + i).border = border

    def create_header_row(sheet, row_num):

        # Fill the values into the cells
        sheet[f'A{row_num}'] = '#'
        sheet[f'B{row_num}'] = 'Channel Information'
        sheet[f'D{row_num}'] = 'Injection Value'
        sheet[f'E{row_num}'] = 'Lower Limit'
        sheet[f'F{row_num}'] = 'Measurement Value'
        sheet[f'G{row_num}'] = 'Upper Limit'
        sheet[f'H{row_num}'] = 'Result'
        sheet[f'I{row_num}'] = 'Notes'

        # Format the cells
        cols = ['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I']
        for col in cols:     
            sheet[f'{col}{row_num}'].style = '20 % - Accent1'
            sheet[f'{col}{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'{col}{row_num}'].border = thin_border

        sheet.merge_cells(f'B{row_num}:C{row_num}')
        sheet.column_dimensions['I'].width = 20

        return row_num + 1

    def create_group_row(sheet, group, row_num):

        # Define some constants for the formatting
        START_COL = 1
        END_COL = 9

        # Merge all the cells in the row for an full divider
        sheet.merge_cells(f'A{row_num}:I{row_num}')

        # Create and format the row for the Group
        sheet[f'A{row_num}'].style = '20 % - Accent3'
        sheet[f'A{row_num}'] = group.name
        sheet[f'A{row_num}'].alignment = Alignment(vertical='center')
        sheet[f'A{row_num}'].font = Font(bold=True)

        return row_num + 1

    def create_channel_row(sheet, channel, row_num):

        # Change out the Eng Units type for the Channel's actual units
        error_type = channel.error_type
        if error_type == 'Eng Units':
            error_type = channel.measurement_units

        # Define some constants for the Channel border
        START_COL = 1
        END_COL = 9
        apply_horizontal_border(sheet, TOP_BORDER_TYPE, "double", row_num, START_COL, END_COL)

        # Fill the values into the cells
        sheet[f'A{row_num}'] = channel.id
        sheet[f'B{row_num}'] = f'Name: {channel.name}'
        sheet[f'C{row_num}'] = f'Tolerance: {channel.max_error} {error_type}'
        sheet[f'B{row_num+1}'] = f'Drawing Ref: ___________________'
        sheet[f'C{row_num+1}'] = f'Interface: ____________________'
        sheet[f'B{row_num+2}'] = f'DC Voltage Source: _____________'
        sheet[f'C{row_num+2}'] = f'Cal Due Date: ________________'
        sheet[f'B{row_num+3}'] = f'MDS: _________________________'
        sheet[f'C{row_num+3}'] = f'Rolls-Royce: _________________'

        # Format the cells            
        sheet[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')

        return row_num

    def create_testpoint_row(sheet, channel, testpoint, row_num):

        # Fill the values into the cells
        sheet[f'D{row_num}'] = f'{testpoint.nominal_injection_value} {channel.injection_units}'
        sheet[f'E{row_num}'] = f'{testpoint.lower_limit()} {channel.measurement_units}'
        sheet[f'F{row_num}'] = f'{channel.measurement_units}'
        sheet[f'G{row_num}'] = f'{testpoint.upper_limit()} {channel.measurement_units}'

        # Format the cells
        sheet[f'F{row_num}'].alignment = Alignment(horizontal='right', vertical='center', indent=1)
        sheet[f'F{row_num}'].border = thin_border
        cols = ['D', 'E', 'G', 'H']
        for col in cols:
            sheet[f'{col}{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'{col}{row_num}'].border = thin_border   

        return row_num + 1

    # Setup the workbook to prepare for the report
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Report'
    sheet = wb[sheet.title]

    # Constants
    NUM_CHANNEL_HEADER_ROWS = 5
    TOP_BORDER_TYPE = 1
    BOTTOM_BORDER_TYPE = 2

    # Setup some initial formatting for the report
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_num = 1
    row_num = create_header_row(sheet, row_num)    

    # Iterate through each Group and its Channels within the Job
    job = Job.query.filter_by(id=job_id).first()
    for group in job.groups.all():
        row_num = create_group_row(sheet, group, row_num)

        for channel in group.channels.all():
            row_num = create_channel_row(sheet, channel, row_num)
            
            # Column formatting
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            sheet['A1'].font = Font(bold=True)

            # Starting row
            start_row = row_num

            for testpoint in channel.testpoints.all():
                row_num = create_testpoint_row(sheet, channel, testpoint, row_num)

            # Set the row for the next Channel if the number of TestPoints interferes with the Channel Information
            num_testpoints = channel.num_testpoints()
            if num_testpoints < NUM_CHANNEL_HEADER_ROWS:
                row_num = start_row + NUM_CHANNEL_HEADER_ROWS

    # Save the new report to the static directory
    directory = current_app.config["TMP_DIRECTORY"]
    filename = 'Job_Report_' + datetime.now().strftime("%m-%d-%Y_%H%M%S") + '.xlsx'
    wb.save(directory + filename)

    # Send the report back to the user
    try:
        return send_from_directory(directory=directory, filename=filename, as_attachment=True)
    except FileNotFoundError:
        return render_template('errors/404.html'), 404